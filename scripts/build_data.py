#!/usr/bin/env python3
"""Compile uploaded court coverage, Census counties, and Tyler markets."""

from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
ROOT = APP_DIR.parent
COURTS_PATH = ROOT / "Coverage map - all_courts.csv"
MARKETS_PATH = ROOT / "Integrator_Markets_5-20.xlsx"
COUNTIES_PATH = APP_DIR / "static" / "data" / "counties.geojson"
OUTPUT_PATH = APP_DIR / "static" / "data" / "coverage.json"

STATE_ABBREVIATIONS = {
    "Alabama": "AL",
    "Alaska": "AK",
    "Arizona": "AZ",
    "Arkansas": "AR",
    "California": "CA",
    "Colorado": "CO",
    "Connecticut": "CT",
    "Delaware": "DE",
    "District of Columbia": "DC",
    "Florida": "FL",
    "Georgia": "GA",
    "Hawaii": "HI",
    "Idaho": "ID",
    "Illinois": "IL",
    "Indiana": "IN",
    "Iowa": "IA",
    "Kansas": "KS",
    "Kentucky": "KY",
    "Louisiana": "LA",
    "Maine": "ME",
    "Maryland": "MD",
    "Massachusetts": "MA",
    "Michigan": "MI",
    "Minnesota": "MN",
    "Mississippi": "MS",
    "Missouri": "MO",
    "Montana": "MT",
    "Nebraska": "NE",
    "Nevada": "NV",
    "New Hampshire": "NH",
    "New Jersey": "NJ",
    "New Mexico": "NM",
    "New York": "NY",
    "North Carolina": "NC",
    "North Dakota": "ND",
    "Ohio": "OH",
    "Oklahoma": "OK",
    "Oregon": "OR",
    "Pennsylvania": "PA",
    "Rhode Island": "RI",
    "South Carolina": "SC",
    "South Dakota": "SD",
    "Tennessee": "TN",
    "Texas": "TX",
    "Utah": "UT",
    "Vermont": "VT",
    "Virginia": "VA",
    "Washington": "WA",
    "West Virginia": "WV",
    "Wisconsin": "WI",
    "Wyoming": "WY",
    "Puerto Rico": "PR",
    "Guam": "GU",
    "American Samoa": "AS",
    "Commonwealth of the Northern Mariana Islands": "MP",
    "U.S. Virgin Islands": "VI",
}

STATE_NAME_FIXES = {
    "Virgina": "Virginia",
    "Washingon": "Washington",
}

APPELLATE_MARKERS = (
    "supreme court",
    "court of appeals",
    "court of appeal",
    "appellate court",
    "appeals court",
    "appellate division",
    "appellate term",
    "tax court",
    "court of claims",
    "court of criminal appeals",
    "court of civil appeals",
)

LOCAL_COURT_MARKERS = (
    "circuit court",
    "circuit and superior courts",
    "district court",
    "superior court",
    "county court",
    "probate court",
    "probate and family court",
    "municipal court",
    "magistrate court",
    "justice court",
    "family court",
    "general sessions court",
    "chancery court",
    "common pleas court",
    "court of common pleas",
    "juvenile court",
)

PRODUCT_NOTES = {
    "CA": "Recommended first certification market. Full EFSP, ECF5 only; Tyler and jurisdiction approvals remain external gates.",
    "IL": "Open market. Confirm whether court scheduling is required for the selected filing context.",
    "IN": "Open market and a plausible post-certification expansion candidate.",
    "MD": "Open market. Administrative certification can be jurisdiction-controlled.",
    "NV": "Open market in the market workbook; validate exact court and filing-family availability.",
    "TX": "Open market with mature filing coverage and separate market/payment complexity.",
}


def normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii").lower()
    value = value.replace("&", " and ")
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def serialize_cell(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def is_tyler_manager(manager: str) -> bool:
    value = normalize(manager)
    signals = (
        "tyler technologies",
        "odyssey",
        "efiletexas",
        "efileil by tyler",
        "enterprise justice file and serve",
        "file and serve tyler technologies",
    )
    return any(signal in value for signal in signals)


def network_for(manager: str, supports_efiling: bool) -> str:
    value = normalize(manager)
    if not supports_efiling or value in {"", "n a"}:
        return "No e-filing"
    if "cm ecf" in value:
        return "Federal CM/ECF"
    if is_tyler_manager(manager):
        return "Tyler / Odyssey"
    if "court approved electronic filing service providers" in value:
        return "EFSP marketplace"
    return "State or local system"


def load_markets() -> dict[str, dict]:
    workbook = load_workbook(MARKETS_PATH, read_only=True, data_only=True)
    markets: dict[str, dict] = {}

    info_sheet = workbook["Market Information"]
    rows = info_sheet.iter_rows(values_only=True)
    next(rows)
    for row in rows:
        raw_name = str(row[0]).strip() if row[0] else ""
        if not raw_name or raw_name.startswith("*"):
            continue
        state_name = STATE_NAME_FIXES.get(raw_name, raw_name)
        state = STATE_ABBREVIATIONS.get(state_name)
        if not state:
            continue
        status = str(row[1]).replace("*", "").strip() if row[1] else "Unknown"
        markets[state] = {
            "state": state,
            "state_name": state_name,
            "status": status,
            "stage_url": str(row[2]).strip() if row[2] else None,
            "production_url": str(row[3]).strip() if row[3] else None,
            "deployments": [],
        }

    update_sheet = workbook["Market Update"]
    rows = update_sheet.iter_rows(values_only=True)
    next(rows)
    for row in rows:
        if not row[0]:
            continue
        state_name = STATE_NAME_FIXES.get(str(row[0]).strip(), str(row[0]).strip())
        state = STATE_ABBREVIATIONS.get(state_name)
        if not state:
            continue
        market = markets.setdefault(
            state,
            {
                "state": state,
                "state_name": state_name,
                "status": "Unknown",
                "stage_url": None,
                "production_url": None,
                "deployments": [],
            },
        )
        market["deployments"].append(
            {
                "category": serialize_cell(row[1]),
                "realm": serialize_cell(row[2]),
                "version": serialize_cell(row[3]),
                "start": serialize_cell(row[4]),
                "status": serialize_cell(row[5]),
            }
        )

    return markets


def main() -> None:
    geography = json.loads(COUNTIES_PATH.read_text())
    county_features = geography["features"]
    counties_by_state: dict[str, list[dict]] = defaultdict(list)
    county_base_counts: dict[str, Counter] = defaultdict(Counter)

    for feature in county_features:
        props = feature["properties"]
        state = props["STUSPS"]
        counties_by_state[state].append(props)
        county_base_counts[state][normalize(props["NAME"])] += 1

    with COURTS_PATH.open(newline="", encoding="utf-8-sig") as handle:
        raw_courts = list(csv.DictReader(handle))

    markets = load_markets()
    court_records = []
    state_records: dict[str, dict[str, list[int]]] = defaultdict(
        lambda: {
            "all": [],
            "federal": [],
            "appellate": [],
            "unresolved": [],
        }
    )
    county_court_ids: dict[str, list[int]] = defaultdict(list)
    county_confidence: dict[str, Counter] = defaultdict(Counter)

    for court_id, row in enumerate(raw_courts):
        court_name = row["Court"].strip()
        raw_state = row["State"].strip()
        states = [part.strip() for part in raw_state.split(",") if len(part.strip()) == 2]
        supports = row["Supports e-filing"].strip().lower() == "yes"
        manager = row["E-filing manager"].strip()
        normalized_name = normalize(court_name)
        combined_text = normalize(f"{court_name} {manager}")

        federal = (
            court_name.startswith("U.S. ")
            or raw_state == "Federal"
            or len(states) > 1
        )
        appellate = not federal and any(
            marker in normalized_name for marker in APPELLATE_MARKERS
        )
        matches: list[dict] = []

        if not federal and not appellate and raw_state in counties_by_state:
            state_name = normalize(
                next(
                    (
                        props["STATE_NAME"]
                        for props in counties_by_state[raw_state]
                    ),
                    raw_state,
                )
            )
            for props in counties_by_state[raw_state]:
                base = normalize(props["NAME"])
                full = normalize(props["NAMELSAD"])
                confidence = None

                if re.search(rf"(?<!\w){re.escape(full)}(?!\w)", combined_text):
                    confidence = "exact-name"
                elif any(
                    f"{descriptor} of {base}" in combined_text
                    for descriptor in (
                        "county",
                        "parish",
                        "borough",
                        "census area",
                        "municipio",
                        "municipality",
                        "city and borough",
                    )
                ):
                    confidence = "exact-name"
                elif raw_state == "DC":
                    confidence = "state-equivalent"
                elif (
                    county_base_counts[raw_state][base] == 1
                    and base != state_name
                    and len(base) >= 3
                    and re.search(rf"(?<!\w){re.escape(base)}(?!\w)", combined_text)
                    and any(marker in normalized_name for marker in LOCAL_COURT_MARKERS)
                ):
                    confidence = "name-inferred"

                if confidence:
                    matches.append(
                        {
                            "geoid": props["GEOID"],
                            "confidence": confidence,
                        }
                    )

        if federal:
            scope = "federal"
        elif appellate:
            scope = "appellate-or-statewide"
        elif matches:
            scope = "county-linked"
        else:
            scope = "state-record-unresolved"

        record = {
            "id": court_id,
            "name": court_name,
            "state_label": raw_state,
            "states": states,
            "supports_efiling": supports,
            "manager": manager,
            "network": network_for(manager, supports),
            "tyler_signaled": is_tyler_manager(manager),
            "scope": scope,
            "county_matches": matches,
        }
        court_records.append(record)

        for state in states:
            state_records[state]["all"].append(court_id)
            if federal:
                state_records[state]["federal"].append(court_id)
            elif appellate:
                state_records[state]["appellate"].append(court_id)
            elif not matches:
                state_records[state]["unresolved"].append(court_id)

        for match in matches:
            geoid = match["geoid"]
            county_court_ids[geoid].append(court_id)
            county_confidence[geoid][match["confidence"]] += 1

    county_records = {}
    availability_counts = Counter()
    counties_with_tyler = 0

    for feature in county_features:
        props = feature["properties"]
        geoid = props["GEOID"]
        court_ids = county_court_ids[geoid]
        linked_courts = [court_records[court_id] for court_id in court_ids]
        yes_count = sum(court["supports_efiling"] for court in linked_courts)
        no_count = len(linked_courts) - yes_count
        manager_counts = Counter(court["manager"] for court in linked_courts)
        network_counts = Counter(court["network"] for court in linked_courts)
        tyler_count = sum(court["tyler_signaled"] for court in linked_courts)

        if not linked_courts:
            availability = "unmapped"
        elif yes_count and no_count:
            availability = "mixed"
        elif yes_count:
            availability = "available"
        else:
            availability = "unavailable"

        availability_counts[availability] += 1
        if tyler_count:
            counties_with_tyler += 1

        state = props["STUSPS"]
        market = markets.get(state)
        if state == "CA":
            rollout = "pilot"
        elif market and market["status"] == "Open":
            rollout = "expansion-candidate"
        else:
            rollout = "not-planned"

        county_records[geoid] = {
            "geoid": geoid,
            "name": props["NAME"],
            "full_name": props["NAMELSAD"],
            "state": state,
            "state_name": props["STATE_NAME"],
            "lsad": props["LSAD"],
            "court_ids": court_ids,
            "efiling_yes": yes_count,
            "efiling_no": no_count,
            "availability": availability,
            "tyler_court_count": tyler_count,
            "networks": dict(network_counts),
            "managers": [
                {
                    "name": manager,
                    "count": count,
                    "tyler_signaled": is_tyler_manager(manager),
                }
                for manager, count in manager_counts.most_common()
            ],
            "match_confidence": dict(county_confidence[geoid]),
            "market_status": market["status"] if market else None,
            "rollout": rollout,
        }

    state_summaries = {}
    for state, county_props in counties_by_state.items():
        county_ids = [props["GEOID"] for props in county_props]
        state_summaries[state] = {
            "state": state,
            "state_name": county_props[0]["STATE_NAME"],
            "county_count": len(county_ids),
            "county_geoids": county_ids,
            "court_ids": state_records[state]["all"],
            "federal_court_ids": state_records[state]["federal"],
            "appellate_court_ids": state_records[state]["appellate"],
            "unresolved_court_ids": state_records[state]["unresolved"],
            "market": markets.get(state),
            "product_note": PRODUCT_NOTES.get(state),
        }

    court_counts = Counter(
        "efile" if record["supports_efiling"] else "no_efile"
        for record in court_records
    )
    scope_counts = Counter(record["scope"] for record in court_records)
    network_counts = Counter(record["network"] for record in court_records)

    payload = {
        "meta": {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "snapshot_date": "2026-06-08",
            "county_vintage": "2025-01-01",
            "counts": {
                "counties": len(county_records),
                "courts": len(court_records),
                "courts_supporting_efiling": court_counts["efile"],
                "courts_without_efiling": court_counts["no_efile"],
                "counties_with_direct_court_matches": sum(
                    bool(record["court_ids"]) for record in county_records.values()
                ),
                "counties_with_tyler_signals": counties_with_tyler,
                "tyler_markets": len(markets),
            },
            "availability_counts": dict(availability_counts),
            "scope_counts": dict(scope_counts),
            "network_counts": dict(network_counts),
            "method": {
                "exact": "County or county-equivalent name appears in the court name or manager field.",
                "inferred": "A unique county basename appears with a local court-type marker.",
                "unresolved": "The source court record does not contain enough geography to assign it to a county without an external district or venue crosswalk.",
            },
        },
        "product": {
            "architecture": "Rust/Axum public service with a private Tyler ECF5 adapter.",
            "certification_target": "CA",
            "certification_status": "Not certified; external approvals and Stage proof remain blocked.",
            "strategy": "Full EFSP, ECF5 only, one California market first.",
            "blocked_signoffs": [
                "Named Tyler BIS certification approver",
                "California administrative certification owner",
                "California Stage and production environment owner",
                "Tyler identity/security owner",
                "Tyler payment owner and platform security owner",
            ],
        },
        "sources": [
            {
                "name": "All courts coverage upload",
                "file": "../Coverage map - all_courts.csv",
                "role": "Court names, state labels, e-filing availability, and managers",
            },
            {
                "name": "Tyler integrator markets",
                "file": "../Integrator_Markets_5-20.xlsx",
                "role": "Market access status, endpoints, and EFM deployment snapshots",
            },
            {
                "name": "U.S. Census Bureau cartographic boundaries",
                "url": "https://www.census.gov/geographies/mapping-files/time-series/geo/cartographic-boundary.2025.html",
                "role": "January 1, 2025 counties and county equivalents",
            },
            {
                "name": "Tyler certification and API dossier",
                "file": "../api-spec/Tyler_Certification_Mapping.md",
                "role": "Certification target, product scope, owners, and unresolved gates",
            },
        ],
        "markets": markets,
        "states": state_summaries,
        "counties": county_records,
        "courts": court_records,
    }

    OUTPUT_PATH.write_text(
        json.dumps(payload, separators=(",", ":"), ensure_ascii=True)
    )
    print(f"Wrote {OUTPUT_PATH}")
    print(json.dumps(payload["meta"]["counts"], indent=2))
    print(json.dumps(payload["meta"]["scope_counts"], indent=2))


if __name__ == "__main__":
    main()
